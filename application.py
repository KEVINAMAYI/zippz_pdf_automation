import math
import datetime
import os
import shutil
import sass
import logging
import json
import requests  # handing api requests
import hyperlink  # formatting links
from openpyxl import load_workbook
from flask import Flask, request, abort, jsonify
from weasyprint import HTML, CSS
from jinja2 import FileSystemLoader
from jinja2 import Environment, select_autoescape
from s3client import upload_files_to_aws, generate_presigned_urls

application = Flask(__name__)
TEMPLATE_NAME = "full"
DELETE_HTML_AFTER_PROCESSING = False
TEST_MODE = False
api_url_for_creating_order = "https://ssapi.shipstation.com/orders/createorder"

SHIPMENTS_FILE = "./files/shipments.xlsx"
INGREDIENTS_FILE_PATH = "./files/ingredients_colors.xlsx"
PROCESSED_FILE_PATH = "./prod/shipments_processed.xlsx"

MIN_ROW = 114
MAX_ROW = 114

logging.basicConfig(level=logging.INFO)

if not os.path.exists("temp"):
    os.makedirs("temp")

env = Environment(
    loader=FileSystemLoader(searchpath="templates"),
    autoescape=select_autoescape(['html', 'xml'])
)

name_mapping = {
    "sleep10": "SleepZ S10",
    "sleep12": "SleepZ S12",
    "sleep14": "SleepZ S14",
    "sleep16": "SleepZ S16",
    "sleep18": "SleepZ S18",
    "sleep20": "SleepZ S20",
    "calm30": "CalmZ C30",
    "calm32": "CalmZ C32",
    "calm34": "CalmZ C34",
    "calm36": "CalmZ C36",
    "calm38": "CalmZ C38",
    "calm40": "CalmZ C40",
}

items_numbers = {
    "calm30": "30-CLM-SG-60",
    "calm32": "32-CLM-SG-60",
    "calm34": "34-CLM-SG-60",
    "calm36": "36-CLM-SG-60",
    "calm38": "38-CLM-SG-60",
    "calm40": "40-CLM-SG-60",
    "sleep10": "10-SLP-SG-60",
    "sleep12": "12-SLP-SG-60",
    "sleep14": "14-SLP-SG-60",
    "sleep16": "16-SLP-SG-60",
    "sleep18": "18-SLP-SG-60",
    "sleep20": "20-SLP-SG-60"
}


def render_html(tempate_name, template_data):
    template = env.get_template("{}.html".format(tempate_name))
    path = "temp/{}".format(template_data['uuid'])
    if (not os.path.exists(path)):
        os.mkdir(path)
    template.stream(template_data).dump('temp/{}/{}.html'.format(template_data['uuid'], tempate_name))


def gen_pdf(template_name: object, template_data: object) -> object:
    render_html(template_name, template_data)
    css = CSS(
        string='')
    if (template_name == "inserts"):
        if (len(template_data.get("issues")) > 1):
            css = CSS(
                string='''
             @page:nth(1) { 
                 size: 576px 384px; 
                 margin: 0; 
             }

            @page:nth(2) { 
                size: 576px 384px; 
                margin: 0; 
                }

            @page:nth(3) { 
                size: 528px 816px; 
                margin: 0; 
            }'''
            )
        else:
            css = CSS(
                string='''
                         @page:nth(1) { 
                             size: 576px 384px; 
                             margin: 0; 
                         }

                        @page:nth(2) { 
                            size: 528px 816px; 
                            margin: 0; 
                        }'''
            )

    uuid = template_data['uuid']
    html = HTML("temp/{}/{}.html".format(uuid, template_name))
    pdf_path = "temp/{}/{}.pdf".format(uuid, template_name)
    html.write_pdf(
        pdf_path, stylesheets=[css])

    if (TEST_MODE == False):
        os.remove(("temp/{}/{}.html".format(uuid, template_name)))
    return pdf_path


def write_signed_urls_to_shippments_file(signed_dict):
    wb = load_workbook(filename=SHIPMENTS_FILE)
    sheet = wb['Orders']
    for row in sheet.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW):
        if (row[1].value is None):
            continue

        uuid = row[9].value;

        if (uuid in signed_dict):
            signed_urls = signed_dict[uuid]
            row[18].value = signed_urls['cards_signed_url']

    wb.save(filename=PROCESSED_FILE_PATH)


def parse_ingredients_legend():
    wb = load_workbook(filename=INGREDIENTS_FILE_PATH)
    sheet = wb["all"]
    names_map = {}
    items = {
        "calm30": [],
        "calm32": [],
        "calm34": [],
        "calm36": [],
        "calm38": [],
        "calm40": [],
        "sleep10": [],
        "sleep12": [],
        "sleep14": [],
        "sleep16": [],
        "sleep18": [],
        "sleep20": [],
        "all": []
    }

    for row in sheet.iter_rows(min_row=2, max_row=2):

        for i in range(3, 15):
            names_map[i] = row[i].value

    for row in sheet.iter_rows(min_row=3, max_row=35):
        color = row[1].value
        name = row[2].value.lower()
        pair = {"color": color, "name": name}
        items["all"].append(pair)
        for i in range(3, 15):

            if (row[i].value == None):
                continue
            product = names_map[i]
            items[product].append(pair)

    return items


# returns a list of customer's data => A customer that has just made a purchase order
def parse_shippments_items(data):
    list = []
    packs = []
    calmz = []
    sleepz = []
    sku = []

    # format sleepz and calmz strings
    def format_string(prodvalue):
        if prodvalue is not None:
            product = prodvalue.replace(" ", "")
            spos = product.index('Z') + 1
            product = product.replace(product[spos], "")
            product = product.replace(product[spos - 1], "")
        return product

    for row in data['line_items']:
        if 'Trial' in row['name']:
            packs.append(row['name'])
            sku.append(row['sku'])
        if 'CalmZ' in row['name']:
            calmz.append(row['name'])
        if 'SleepZ' in row['name']:
            sleepz.append(row['name'])

    fname = data['billing']['first_name']
    lname = data['billing']['last_name']
    email = data['billing']['first_name']
    street1 = data['billing']['address_1']
    street2 = data['billing']['address_2']
    city = data['billing']['city']
    state = data['billing']['state']
    zipcode = data['billing']['postcode']
    pack = packs[0]
    sleep1 = format_string(sleepz[0]) if 0 < len(sleepz) else None
    sleep2 = format_string(sleepz[1]) if 1 < len(sleepz) else None
    calm1 = format_string(calmz[0]) if 0 < len(calmz) else None
    calm2 = format_string(calmz[1]) if 1 < len(calmz) else None
    orderno = data['id']
    lineitems = data['line_items']
    customerid = data['customer_id']
    orderkey = data['order_key']
    createdate = data['date_created']
    country = data['billing']['country']
    date_modified = data['date_modified']
    created = datetime.datetime.strptime(createdate, '%Y-%m-%dT%H:%M:%S')
    orderdate = created.strftime("%m/%d/%Y")
    datetitle = created.strftime('%B %d, %Y')
    data = '{"action":"create", "id":"' + pack + '"}'
    items = [{
        "orderItemId": lineitems[0]['id'],
        "lineItemKey": None,
        "sku": sku[0],
        "name": lineitems[0]['name'],
        "imageUrl": None,
        "weight": None,
        "quantity": lineitems[0]['quantity'],
        "unitPrice": lineitems[0]['subtotal'],
        "taxAmount": None,
        "shippingAmount": None,
        "warehouseLocation": None,
        "options": [],
        "productId": lineitems[0]['product_id'],
        "fulfillmentSku": None,
        "adjustment": False,
        "upc": None,
        "createDate": createdate,
        "modifyDate": date_modified
    }]

    customer = {
        'first': fname,
        'last': lname,
        'email': email,
        'street1': street1,
        'street2': street2,
        'city': city,
        'state': state,
        'zip': zipcode,
        'pack': pack,
        'sleep1': sleep1,
        'sleep2': sleep2,
        'calm1': calm1,
        'calm2': calm2,
        'uuid': orderno,
        'order_number': orderno,
        'date_order': orderdate,
        'date_title': datetitle,

    }
    list.append(customer)

    # create a dictionary to store order data for passing to ship station
    parse_shippments_items.order_for_shipstation = {
        "orderNumber": orderno,
        "orderKey": orderno,
        "orderDate": orderdate,
        "paymentDate": orderdate,
        "shipByDate": None,
        "orderStatus": "awaiting_shipment",
        "customerId": customerid,
        "customerUsername": None,
        "customerEmail": None,
        "billTo": {
            "name": fname + " " + lname,
            "company": None,
            "street1": street1,
            "street2": street2,
            "street3": None,
            "city": city,
            "state": state,
            "postalCode": zipcode,
            "country": country,
            "phone": None,
            "residential": None
        },
        "shipTo": {
            "name": fname + " " + lname,
            "company": None,
            "street1": street1,
            "street2": street2,
            "street3": None,
            "city": city,
            "state": state,
            "postalCode": zipcode,
            "country": country,
            "phone": None,
            "residential": None
        },
        "items": items,
        "amountPaid": 0.0,
        "taxAmount": 0.0,
        "shippingAmount": 0.0,
        "customerNotes": "Please ship the product as as soon as possible!",
        "internalNotes": None,
        "gift": False,
        "giftMessage": None,
        "paymentMethod": None,
        "requestedShippingService": None,
        "carrierCode": "stamps_com",
        "serviceCode": "usps_first_class_mail",
        "packageCode": "package",
        "confirmation": "delivery",
        "shipDate": None,
        "weight": {
            "value": 15.00,
            "units": "ounces",
            "WeightUnits": 1
        },
        "dimensions": {
            "units": "inches",
            "length": 5.00,
            "width": 7.00,
            "height": 3.00
        },
        "insuranceOptions": {
            "provider": None,
            "insureShipment": False,
            "insuredValue": 0.0
        },
        "internationalOptions": {
            "contents": None,
            "customsItems": None
        },
        "advancedOptions": {
            "warehouseId": 638301,
            "nonMachinable": False,
            "saturdayDelivery": False,
            "containsAlcohol": False,
            "mergedOrSplit": False,
            "mergedIds": [],
            "parentId": None,
            "storeId": 300914,
            "customField1": None,
            "customField2": None,
            "customField3": None,
            "source": "None",
            "billToParty": None,
            "billToAccount": None,
            "billToPostalCode": None,
            "billToCountryCode": None
        },
        "tagIds": None,
        "userId": "50ad70dd-4b1f-421a-bb4f-39c792d41690",
        "externallyFulfilled": False,
        "externallyFulfilledBy": None,
    }

    return list


# return specific products with their benefits
def parse_ingredients():
    dict = {}
    wb = load_workbook(filename='files/ingredients.xlsx')
    sheet = wb["SleepZ"]
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        if row[1].value is None:
            continue
        key = row[1].value
        key = key.lower().replace(" ", "").replace("z", "")

        dict[key] = [
            row[4].value,
            row[5].value,
            row[6].value
        ]
    sheet = wb["CalmZ"]
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        if row[1].value is None:
            continue
        key = row[1].value
        key = key.lower().replace(" ", "").replace("z", "")

        dict[key] = [
            row[3].value,
            row[4].value,
            row[5].value
        ]

    return dict


def my_key(x):
    if x['name'] == "cbd":
        return "0"
    if x['name'] == "cbg":
        return "1"
    if x['name'] == "cbn":
        return "2"
    else:
        return x['name']


def compile_scss():
    sass.compile(dirname=('sass', 'css'))


def delete_test_data():
    if os.path.exists("temp/test"):
        shutil.rmtree("temp/test")


def generate_pdfs_for_shippment(customer, ingredients, ingredients_legend):
    uuid = customer['uuid']
    if TEST_MODE == True:
        uuid = "test"

    data = {
        "uuid": uuid,
        "order_number": customer['order_number'],
        "email": customer['email'],
        "first": customer['first'],
        "last": customer['last'],
        "street1": customer['street1'],
        "street2": customer['street2'],
        "state": customer['state'],
        "city": customer['city'],
        "zip": customer['zip'],
        "date_title": customer['date_title'],
        "date_order": customer['date_order'],
    }

    issues = []
    legend = []

    if (customer['calm1'] is not None
            and customer['calm2'] is not None
            and customer['calm1'][0] != "x"):
        calms = [customer['calm1'].lower(), customer['calm2'].lower()]
        calms.sort()

        calm1_id = calms[0]
        calm2_id = calms[1]

        legend += ingredients_legend[calm1_id]
        legend += ingredients_legend[calm2_id]

        calm1_benefits = ingredients[calm1_id]
        calm2_benefits = ingredients[calm2_id]
        calm1_title = name_mapping[calm1_id]
        calm2_title = name_mapping[calm2_id]
        calm1_sku = items_numbers[calm1_id]
        calm2_sku = items_numbers[calm2_id]

        issue = {
            "type": "calmz",
            "product1": {"name": calm1_title, "id": calm1_id, "sku": calm1_sku, "benefits": calm1_benefits},
            "product2": {"name": calm2_title, "id": calm2_id, "sku": calm2_sku, "benefits": calm2_benefits},
            "instructions": generate_instructions("calmz", calm1_title, calm2_title),
            "faq_instructions": generate_faq_instructions("calmz", calm1_title, calm2_title)
        }
        issues.append(issue)

    if (customer['sleep1'] is not None
            and customer['sleep2'] is not None
            and customer['sleep1'][0] != "x"):
        sleeps = [customer['sleep1'].lower(), customer['sleep2'].lower()]
        sleeps.sort()

        sleep1_id = sleeps[0]
        sleep2_id = sleeps[1]

        legend += ingredients_legend[sleep1_id]
        legend += ingredients_legend[sleep2_id]

        sleep1_benefits = ingredients[sleep1_id]
        sleep2_benefits = ingredients[sleep2_id]
        sleep1_title = name_mapping[sleep1_id]
        sleep2_title = name_mapping[sleep2_id]
        sleep1_sku = items_numbers[sleep1_id]
        sleep2_sku = items_numbers[sleep2_id]

        issue = {
            "type": "sleepz",
            "product1": {"name": sleep1_title, "id": sleep1_id, "sku": sleep1_sku, "benefits": sleep1_benefits},
            "product2": {"name": sleep2_title, "id": sleep2_id, "sku": sleep2_sku, "benefits": sleep2_benefits},
            "instructions": generate_instructions("sleepz", sleep1_title, sleep2_title),
            "faq_instructions": generate_faq_instructions("sleepz", sleep1_title, sleep2_title)
        }
        issues.append(issue)

    legend_without_duplicates = [dict(t) for t in {tuple(d.items()) for d in legend}]
    legend_without_duplicates = sorted(legend_without_duplicates, key=my_key)

    data['issues'] = issues

    arr_len = len(legend_without_duplicates)
    column_size = math.ceil(arr_len / 2)
    data['legend_column1'] = legend_without_duplicates[:column_size]
    data['legend_column2'] = legend_without_duplicates[column_size:33]

    inserts_path = "/test"
    inserts_path = ""
    cards_path = gen_pdf("cards", data)
    return inserts_path, cards_path


def generate_faq_instructions(type, product1, product2):
    instructions = []
    if (type == "calmz"):
        instruction1 = "It’s ok to skip some days, or not finish a bottle if it isn’t working for you."

        instruction2 = "You can also take CalmZ multiple times per day."

        instructions.extend([instruction1, instruction2]);

    elif (type == "sleepz"):
        instruction1 = "It’s ok to skip some days, or not finish a bottle if it isn’t working for you."

        instructions.extend([instruction1]);
    return instructions


def generate_instructions(type, product1, product2):
    instructions = []
    if (type == "calmz"):
        instruction1 = {
            "bold_text": "Start with {}".format(product1),
            "text": "Take each day for 5 days."
        }

        instruction2 = {
            "bold_text": "Switch to {}".format(product2),
            "text": "Take each day for 5 days."
        }

        instruction3 = {
            "bold_text": "Suggested Use",
            "text": "Take 2 softgels as needed. Best when taken with food. Allow 4-5 hours between doses."
        }

        instruction4 = {
            "bold_text": "Track Your Progress",
            "text": "Use the rating card after each dose."
        }

        instructions.extend([instruction1, instruction2, instruction3, instruction4]);

    elif type == "sleepz":
        instruction1 = {
            "bold_text": "Start with {}".format(product1),
            "text": "Take each night for 5 nights."
        }

        instruction2 = {
            "bold_text": "Switch to {}".format(product2),
            "text": "Take each night for 5 nights."
        }

        instruction3 = {
            "bold_text": "Suggested Use",
            "text": "Take 2 softgels 30-60 minutes before bedtime."
        }

        instruction4 = {
            "bold_text": "Track Your Progress",
            "text": "Use the rating card after each dose."
        }

        instructions.extend([instruction1, instruction2, instruction3, instruction4]);
    return instructions


# function to shorten the AWS S3 url using rebrandly apis
def shorten_url(pdf_url):
    link_request = {
        "destination": pdf_url,
        "domain": {"fullName": "rebrand.ly"}
    }

    # header parameters
    request_headers = {
        "Content-type": "application/json",
        "apikey": "bf0a166016c74d17afaffdd1656d9bef",
        "workspace": "1562030ed25a4759b1922a5150e438ad"
    }

    # get resonse
    r = requests.post("https://api.rebrandly.com/v1/links",
                      data=json.dumps(link_request),
                      headers=request_headers)

    # response status and store url in a variable
    if r.status_code == requests.codes.ok:
        link = r.json()
        return link["shortUrl"]


# function for attaching pdf url to order
def attach_pdf_url_to_order(order, order_pdf_url):
    # get the order pdf url passed through the function and format
    url = hyperlink.parse(order_pdf_url)

    # attach the formatted pdf url along with custom text to the customfield1 index of the order
    order["advancedOptions"]["customField1"] = "Testing update for pdf  https://" + url.to_text() + ' '

    # return an order with pdf url at the customfield1 index
    return order


# function for creating/updating orders in shipstation
def create_update_order_in_shipstation(order_data):
    # define the headers for the api
    headers = {
        'Content-Type': 'application/json',
        'Authorization': 'Basic ZGNiMDM5M2ZiNzU2NDQxYWJhZGEyOTA3NTc2YWMwODM6NWZiM2VkNTA0ZGY4NDQ3ODk2ZWZjN2M4OGEwZDI1OTI='
    }

    # post the data to shipstation and check the response
    response = requests.request("POST", api_url_for_creating_order, headers=headers, data=json.dumps(order_data))
    print(response.text)


# expose an endpoint for getting customer data from wordpress
@application.route('/')
def test():
    return "python pdf generation app"


# expose an endpoint for getting customer data from wordpress
@application.route('/api/post/order', methods=["POST"])
def get_order_data_from_wordpress():
    if request.method == 'POST':

        # get order data as a string and convert to json
        order = request.json
        print(order)
        delete_test_data()
        # compile_scss()
        ingredients = parse_ingredients()
        ingredients_legend = parse_ingredients_legend()
        shippments = parse_shippments_items(order)
        signed_urls = {}
        for shippment in shippments:
            uuid = shippment['uuid']
            logging.info("Start generating pdfs for shipments with email {}".format(shippment['email']))
            inserts_path, cards_path = generate_pdfs_for_shippment(shippment, ingredients, ingredients_legend)
            logging.info("Finished generating pdfs for shipments with email {}".format(shippment['email']))
            logging.info("Start uploading pdfs for shipments with email {}".format(shippment['email']))
            if (TEST_MODE == False):
                upload_files_to_aws([inserts_path, cards_path], uuid)
                inserts_signed_url, cards_signed_url = generate_presigned_urls(shippment['uuid'])
                signed_urls[uuid] = {"inserts_signed_url": inserts_signed_url, "cards_signed_url": cards_signed_url}
                logging.info("Finished uploading pdfs for shipments with email {}".format(shippment['email']))

        logging.info("Start writing urls to file")
        logging.info(signed_urls)
        write_signed_urls_to_shippments_file(signed_urls)

        # call the functions for shortening pdf_url, attaching pdf_url to order and send order details to shipstation
        pdf_shortened_url = shorten_url(cards_signed_url)
        order_with_pdf_url = attach_pdf_url_to_order(parse_shippments_items.order_for_shipstation, pdf_shortened_url)
        create_update_order_in_shipstation(order_with_pdf_url)
        return 'success', 200


# runs main file
if __name__ == '__main__':
    application.run(host='127.0.0.1', port=80)
